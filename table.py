import pandas as pd
from datetime import datetime, timedelta 
import numpy as np
import openpyxl

gg = input('no: ')

df = pd.read_excel('SIM2to5Report ({}).xlsx'.format(gg))

df.rename(df.loc[0], axis='columns', inplace=True)

pd.set_option('display.max_rows',86)

df.drop(0, inplace=True)

l = int(input('Week: '))
f1 = 7*(l-1)

k = str(l)
n = 'week_{}'.format(k)
print(n)

df

a = datetime(2021, 1, 3, 23, 59) + timedelta(days=f1)
b = datetime(2021, 1, 10, 0, 0) + timedelta(days=f1)

#filt = (df['Escalate To'] == 'SIM4') & (df['Date Occured'] > '2021-02-15')
filt = (df['Escalate To'] == 'SIM4') & (df['Start Time'] > a) & (df['Start Time'] < b)

x = df.loc[filt, ['ID', 'Owner', 'Start Time']]

y = x['Owner'].value_counts()

dfx = pd.read_excel('Schneider_names.xlsx', index_col='S No')

f = y.to_dict()

z = dfx.to_dict()

y

z[n] = {}

z

kk = [d.lower() for d in z['Name'].values()]

nop1 = {k.lower(): v for k, v in f.items()}

kk

g = 1
for i in kk:
    try:
        z[n][g] = nop1[i]
    except KeyError:
        z[n][g] = 0
    g += 1

z

xfd = pd.DataFrame.from_dict(z)

xfd.index.name = 'S.No'

xfd

xfd.to_excel('Excel_task.xlsx')

# opps = openpyxl.load_workbook('Excel_task.xlsx')
# ws = opps.active

# openpyxl.worksheet.dimensions.ColumnDimension(ws, bestFit=True)

# opps.save('new_doc.xlsx')

wb = openpyxl.load_workbook('Excel_task.xlsx')
ws = wb.active

ws.column_dimensions['B'].width = 27
ws.column_dimensions['C'].width = 22
wb.save("column_width_test.xlsx")

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
wb = openpyxl.load_workbook('column_width_test.xlsx')
ws = wb.active
YelFill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
LredFill = PatternFill(start_color='00FF8080', end_color='00FF8080', fill_type='lightDown')
ws.conditional_formatting.add('D2:D87', CellIsRule(operator='equal', formula=[0], stopIfTrue=True, fill=YelFill))

ws.conditional_formatting.add('D2:D87', CellIsRule(operator='lessThan', formula=[8], stopIfTrue=True, fill=LredFill))
wb.save("Result_test.xlsx")

